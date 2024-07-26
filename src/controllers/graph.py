import io
import locale
import openpyxl

from decimal import Decimal
from datetime import datetime
from flask import abort, request, make_response, Blueprint
from openpyxl.chart import BarChart, Reference
from openpyxl.chart.label import DataLabel, DataLabelList

from Service.generate_sales_graph import generatesalesgraph

locale.setlocale(locale.LC_ALL, 'pt_BR.UTF-8')  # Define a localidade para português brasileiro

REQUEST_API = Blueprint('graph', __name__)

# Função para calcular a altura do gráfico com base no número de itens
def calculate_chart_height(num_items):
    base_height = 15  # Altura base do gráfico
    height_per_item = 2  # Altura adicional por item
    return base_height + height_per_item * (num_items // 5)  # Ajusta a altura dependendo do número de itens

def create_chart_sheet(workbook, data_sheet, start_index, end_index, sheet_index):
    chart_sheet = workbook.create_sheet(title=f'Gráfico {sheet_index}')

    produtos = [row[0] for row in data_sheet.iter_rows(min_row=start_index, max_row=end_index, min_col=1, max_col=1)]
    quantidade = [row[0] for row in data_sheet.iter_rows(min_row=start_index, max_row=end_index, min_col=2, max_col=2)]
    valor_produto = [row[0] for row in data_sheet.iter_rows(min_row=start_index, max_row=end_index, min_col=3, max_col=3)]
    valor_formatado = [row[0] for row in data_sheet.iter_rows(min_row=start_index, max_row=end_index, min_col=4, max_col=4)]

    # Adicionar gráfico de barras para a quantidade
    chart_qtd = BarChart()
    chart_qtd.type = "bar"
    chart_qtd.style = 10

    chart_qtd.title = "Gráfico referente à quantidade de venda"

    categories_ref = Reference(data_sheet, min_col=1, min_row=start_index, max_row=end_index)
    qtd_ref = Reference(data_sheet, min_col=2, min_row=start_index, max_row=end_index)

    chart_qtd.add_data(qtd_ref, titles_from_data=True)
    chart_qtd.set_categories(categories_ref)

    # Centralizar os rótulos no centro da barra
    for series in chart_qtd.series:
        series.dLbls = DataLabelList()
        series.dLbls.showVal = True
        series.dLbls.showCatName = False
        series.dLbls.dLblPos = 'outEnd'  # Posicionar no centro da barra

    chart_qtd.y_axis.majorGridlines = None
    chart_qtd.legend.position = "r"  # Posições possíveis: 'r' (direita), 't' (acima), 'l' (esquerda), 'b' (abaixo)
    chart_qtd.width = 30  # Ajuste conforme necessário
    chart_qtd.height = calculate_chart_height(len(produtos))  # Ajuste conforme necessário
    chart_qtd.barWidth = 10  # Ajustar a largura das barras para criar espaço extra
    chart_qtd.gapWidth = 60  # Ajustar o espaçamento entre as barras

    chart_sheet.add_chart(chart_qtd, "B2")

    # Adicionar gráfico de barras para o valor
    chart_valor = BarChart()
    chart_valor.type = "bar"
    chart_valor.style = 10

    # Adicionar título ao gráfico de valor
    chart_valor.title = "Gráfico referente ao valor do produto"

    valor_ref = Reference(data_sheet, min_col=3, min_row=start_index, max_row=end_index)

    chart_valor.add_data(valor_ref, titles_from_data=False)
    chart_valor.set_categories(categories_ref)

    # Centralizar os rótulos no centro da barra
    for series in chart_valor.series:
        series.dLbls = DataLabelList()
        series.dLbls.showVal = True
        series.dLbls.showCatName = False
        series.dLbls.dLblPos = 'outEnd'  # Posicionar no centro da barra

    chart_valor.y_axis.majorGridlines = None
    chart_valor.legend.position = "r"  # Posições possíveis: 'r' (direita), 't' (acima), 'l' (esquerda), 'b' (abaixo)
    chart_valor.width = 30  # Ajuste conforme necessário
    chart_valor.height = calculate_chart_height(len(produtos))  # Ajuste conforme necessário
    chart_valor.barWidth = 10  # Ajustar a largura das barras para criar espaço extra
    chart_valor.gapWidth = 60  # Ajustar o espaçamento entre as barras

    chart_sheet.add_chart(chart_valor, "W2")  # Ajustar a posição para não sobrepor o gráfico de quantidade

    # Remover grades vazias e ocultar linhas e colunas da planilha de gráficos
    chart_sheet.sheet_view.showGridLines = False

def adjust_column_width(sheet):
    for col in sheet.columns:
        max_length = 0
        column = col[0].column_letter  # Obtém a letra da coluna
        
        for cell in col:
            try:
                if cell.value is not None and len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        
        adjusted_width = max_length + 2  # Adiciona um pouco de espaço extra
        sheet.column_dimensions[column].width = adjusted_width

@REQUEST_API.route('/fetch_sales', methods=['GET'])
def fetch_sales():
    dt_inicial = request.args.get('dt_inicial')
    dt_final = request.args.get('dt_final')
    idstatus = request.args.get('idstatus')
    idcategoria = request.args.get('idcategoria')
    idusuario = request.args.get('idusuario')

    if not dt_inicial and not dt_final and not idstatus and not idcategoria and not idusuario:
        abort(400, description="Parâmetros insuficientes")

    response = generatesalesgraph(dt_inicial, dt_final, idstatus, idcategoria, idusuario)

    produtos = []
    quantidade = []
    valor_produto = []
    valor_formatado = []

    for item in response:
        descricao = item['descricao']
        modelo = item['modelo']
        qtd = int(Decimal(item['qtd']))
        valor = float(Decimal(item['valor']))

        produtos.append(descricao + ' - ' + modelo)
        quantidade.append(qtd)
        valor_produto.append(valor)

        locale.setlocale(locale.LC_ALL, 'pt_BR.UTF-8')
        valor_formatado.append(locale.currency(valor, grouping=True))

    # Criar uma planilha Excel em memória
    workbook = openpyxl.Workbook()

    # Adicionar a aba de dados de vendas
    data_sheet = workbook.create_sheet(title='Dados de Vendas')

    # Escrever o cabeçalho do XLSX
    data_sheet.append(['Produto', 'Quantidade', 'Valor', 'Valor Formatado'])

    # Escrever os dados no XLSX
    for prod, qtd, val, val_fmt in zip(produtos, quantidade, valor_produto, valor_formatado):
        data_sheet.append([prod, qtd, val, val_fmt])

    # Ajustar a largura das colunas na aba de dados
    adjust_column_width(data_sheet)

    # Dividir os dados em partes de 75 itens
    chunk_size = 30
    num_chunks = (len(produtos) + chunk_size - 1) // chunk_size

    for i in range(num_chunks):
        start_index = i * chunk_size + 2  # +2 para pular o cabeçalho
        end_index = min((i + 1) * chunk_size + 1, len(produtos) + 1)  # +1 para incluir o último item

        create_chart_sheet(workbook, data_sheet, start_index, end_index, i + 1)

    # Adicionar a tabela de vendas final na última aba
    final_sheet.append(['Produto', 'Quantidade', 'Valor', 'Valor Formatado'])
    for prod, qtd, val, val_fmt in zip(produtos, quantidade, valor_produto, valor_formatado):
        final_sheet.append([prod, qtd, val, val_fmt])

    # Ajustar a largura das colunas na aba final
    adjust_column_width(final_sheet)

    # Mover a planilha 'Dados de Vendas' para o final
    data_sheet.sheet_view.showGridLines = False
    workbook.move_sheet(data_sheet, offset=-1)  # Move 'Dados de Vendas' para o final

    # Salvar o arquivo XLSX em memória usando BytesIO
    output = io.BytesIO()
    workbook.save(output)
    xlsx_data = output.getvalue()
    output.close()

    # Criar um nome de arquivo com a data atual
    current_date = datetime.now().strftime('%d-%m-%Y')  # Formato: '26-07-2024'
    filename = f'Relatorio-Vendas {current_date}.xlsx'

    # Criar uma resposta Flask para download do arquivo XLSX
    response = make_response(xlsx_data)
    response.headers['Content-Disposition'] = f'attachment; filename={filename}'
    response.headers['Content-type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'

    return response
